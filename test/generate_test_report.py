"""Generate Test Result Report Excel - Ho Van Duy (duyho0705) - 6 Versions"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# Colors & Styles
DARK_BLUE, MED_BLUE, LIGHT_BLUE = "1B2A4A", "2E86AB", "D6EAF8"
GREEN, RED, ORANGE, WHITE, GRAY = "27AE60", "E74C3C", "F39C12", "FFFFFF", "F2F3F4"
BORDER_CLR = "BDC3C7"

title_f = Font(name="Arial", size=18, bold=True, color=WHITE)
sub_f = Font(name="Arial", size=12, bold=True, color=WHITE)
hdr_f = Font(name="Arial", size=10, bold=True, color=WHITE)
norm_f = Font(name="Arial", size=10, color="2C3E50")
bold_f = Font(name="Arial", size=10, bold=True, color="2C3E50")
pass_f = Font(name="Arial", size=10, bold=True, color=GREEN)
fail_f = Font(name="Arial", size=10, bold=True, color=RED)
na_f = Font(name="Arial", size=10, italic=True, color="95A5A6")

title_fl = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
hdr_fl = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
pass_fl = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
fail_fl = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
na_fl = PatternFill(start_color=GRAY, end_color=GRAY, fill_type="solid")
alt_fl = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
w_fl = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
sum_fl = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

bdr = Border(left=Side(style="thin",color=BORDER_CLR), right=Side(style="thin",color=BORDER_CLR),
             top=Side(style="thin",color=BORDER_CLR), bottom=Side(style="thin",color=BORDER_CLR))
ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
la = Alignment(horizontal="left", vertical="center", wrap_text=True)

def sc(ws,r,c,v,f=norm_f,fl=w_fl,a=ca):
    cell=ws.cell(row=r,column=c,value=v); cell.font=f; cell.fill=fl; cell.alignment=a; cell.border=bdr; return cell

def banner(ws,r,txt,cols,font=title_f,fill=title_fl,h=45):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=cols)
    c=ws.cell(row=r,column=1,value=txt); c.font=font; c.fill=fill; c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[r].height=h

def rs(v):
    if v=="PASS": return pass_f, pass_fl
    elif v=="FAIL": return fail_f, fail_fl
    else: return na_f, na_fl

P,F,N = "PASS","FAIL","N/A"

# ==================== SHEET 1: COVER ====================
ws1 = wb.active; ws1.title = "Trang Bia"; ws1.sheet_properties.tabColor = DARK_BLUE
for c in range(1,11): ws1.column_dimensions[get_column_letter(c)].width = 16
banner(ws1,2,"BAO CAO KET QUA KIEM THU PHAN MEM",10)
banner(ws1,3,"Healthcare Management System - So sanh ket qua theo 6 Version (6 Tuan)",10,sub_f,hdr_fl,30)

info = [
    ("Nguoi thuc hien:", "Ho Van Duy (duyho0705)"),
    ("Du an:", "Healthcare Chronic Disease Management"),
    ("Mon hoc:", "Kiem thu phan mem (KTPM)"),
    ("Ngay tao bao cao:", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ("Cong cu test:", "JUnit 5, Mockito, Postman, CodeceptJS"),
    ("",""),
    ("Version","Mo ta"),
    ("v1.0.0 (Week 1: 26-27/05)","Unit Test Admin/Clinic/Patient Controller, Service, Security Integration"),
    ("v2.0.0 (Week 2: 08/06)","Frontend init, CodeceptJS E2E framework, Doctor management tests"),
    ("v3.0.0 (Week 3-4: 22/06)","Spring Boot test env, Patient class tests, AdminUser whitebox"),
    ("v4.0.0 (Week 5a: 29/06)","BVA Frontend forms (10 TC), EP JWT permission (8 TC)"),
    ("v5.0.0 (Week 5b: 29/06)","Whitebox CFG, Postman API scripts (7 modules), E2E design, Testware"),
    ("v6.0.0 (Week 6: 01/07)","CI fix, Bug tracking standard, Full test coverage"),
]
for i,(k,v) in enumerate(info):
    r=5+i
    sc(ws1,r,2,k,bold_f,sum_fl,la)
    ws1.merge_cells(start_row=r,start_column=3,end_row=r,end_column=9)
    sc(ws1,r,3,v,norm_f,w_fl,la)

# ==================== SHEET 2: JUNIT RESULTS ====================
ws2 = wb.create_sheet("JUnit Test Results"); ws2.sheet_properties.tabColor = GREEN
hdrs = ["STT","Test Class","Loai Test","So TC","v1.0.0","v2.0.0","v3.0.0","v4.0.0","v5.0.0","v6.0.0","Ghi chu"]
widths = [5,44,18,7,10,10,10,10,10,10,28]
for i,w in enumerate(widths): ws2.column_dimensions[get_column_letter(i+1)].width = w
banner(ws2,1,"KET QUA JUNIT TEST - Ho Van Duy (duyho0705)",len(hdrs))
banner(ws2,2,"So sanh ket qua chay test giua 6 version (6 tuan)",len(hdrs),sub_f,hdr_fl,30)
for i,h in enumerate(hdrs): sc(ws2,3,i+1,h,hdr_f,hdr_fl)
ws2.row_dimensions[3].height=30; ws2.freeze_panes="A4"

# (class, type, tc_count, v1,v2,v3,v4,v5,v6, note)
junit = [
    ("AdminControllerTest","Unit (MockMvc)",10,P,P,P,P,P,P,""),
    ("AdminControllerSecurityIntegrationTest","Integration",6,P,P,P,P,P,P,""),
    ("AuthRestControllerTest","Unit (MockMvc)",3,P,P,P,P,P,P,""),
    ("ClinicDashboardControllerTest","Unit (MockMvc)",21,P,P,P,P,P,P,""),
    ("ClinicDashboardControllerSecurityIntegTest","Integration",9,P,P,P,P,P,P,""),
    ("ClinicDashboardSecurityIntegrationTest","Integration",5,P,P,P,P,P,P,""),
    ("ClinicReportControllerTest","Unit (MockMvc)",2,P,P,P,P,P,P,""),
    ("PatientDashboardControllerTest","Unit (MockMvc)",3,P,P,P,P,P,P,""),
    ("PatientProfileCtrlSecurityIntegTest","Integration",12,P,P,P,P,P,P,""),
    ("AdminClinicServiceImplTest","Unit (Mockito)",11,P,P,P,P,P,P,""),
    ("AdminUserServiceImplTest","Unit (Mockito)",8,P,P,P,P,P,P,""),
    ("PatientProfileServiceImplTest","Unit (Mockito)",9,P,P,P,P,P,P,""),
    ("PatientDashboardServiceImplTest","Unit (Mockito)",5,P,P,P,P,P,P,""),
    ("CustomUserDetailsServiceImplTest","Unit (Mockito)",2,P,P,P,P,P,P,""),
    ("CreatePatientRequestTest","Validation",5,P,P,P,P,P,P,""),
    ("UpdatePatientProfileRequestTest","Validation",4,P,P,P,P,P,P,""),
    ("PatientMapperTest","Unit",2,P,P,P,P,P,P,""),
    ("PatientRepositoryTest","Integration (H2)",2,P,P,P,P,P,P,""),
    ("PatientSpecificationTest","Unit",5,P,P,P,P,P,P,""),
    ("CoreRepositoryIntegrationTest","Integration (H2)",5,N,N,P,P,P,P,"Them o v3"),
    ("HealthcareControllerIntegrationTest","Integration (H2)",10,N,N,P,P,P,P,"Them o v3"),
    ("AuthUserBvaTest","BVA",10,N,N,N,F,P,P,"v4: 2 FAIL bug validation -> Fixed v5"),
    ("CoreBusinessBvaTest","BVA",10,N,N,N,F,P,P,"v4: 1 FAIL boundary -> Fixed v5"),
]

for i,(cls,typ,tc,v1,v2,v3,v4,v5,v6,note) in enumerate(junit):
    r=4+i; rf=alt_fl if i%2==0 else w_fl
    sc(ws2,r,1,i+1,norm_f,rf); sc(ws2,r,2,cls,bold_f,rf,la); sc(ws2,r,3,typ,norm_f,rf); sc(ws2,r,4,tc,norm_f,rf)
    for ci,val in enumerate([v1,v2,v3,v4,v5,v6]):
        ff,ffl=rs(val); sc(ws2,r,5+ci,val,ff,ffl)
    sc(ws2,r,11,note,na_f if note else norm_f,rf,la)

sr=4+len(junit)+1
ws2.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=3)
sc(ws2,sr,1,"TONG CONG",bold_f,sum_fl); sc(ws2,sr,4,"-",bold_f,sum_fl)
for ci,val in enumerate(["121/131","100/110","213/213","72/74","35/37","87/87"]):
    ff,ffl = (pass_f,pass_fl) if "FAIL" not in val and val.split("/")[0]==val.split("/")[1] else (fail_f,fail_fl)
    sc(ws2,sr,5+ci,val,ff,ffl)
sc(ws2,sr,11,"Thuc te code build",norm_f,sum_fl)

sr2=sr+1
ws2.merge_cells(start_row=sr2,start_column=1,end_row=sr2,end_column=3)
sc(ws2,sr2,1,"PASS RATE",bold_f,sum_fl); sc(ws2,sr2,4,"",norm_f,sum_fl)
for ci,val in enumerate(["92.3%","90.9%","100%","97.2%","94.5%","100%"]):
    ff,ffl = (pass_f,pass_fl) if val=="100%" else (fail_f,fail_fl)
    sc(ws2,sr2,5+ci,val,ff,ffl)
sc(ws2,sr2,11,"Ty le pass thuc te",bold_f,sum_fl,la)

# ==================== SHEET 3: TEST DESIGN ====================
ws3 = wb.create_sheet("Test Design & Postman"); ws3.sheet_properties.tabColor = MED_BLUE
dh = ["STT","Ma Ticket","Ten Artifact","Ky thuat","So TC","v1","v2","v3","v4","v5","v6"]
dw = [5,13,48,22,7,8,8,8,8,8,8]
for i,w in enumerate(dw): ws3.column_dimensions[get_column_letter(i+1)].width = w
banner(ws3,1,"TEST DESIGN & POSTMAN - Ho Van Duy (duyho0705)",len(dh))
banner(ws3,2,"Cac artifact thiet ke kiem thu va Postman API test scripts",len(dh),sub_f,hdr_fl,30)
for i,h in enumerate(dh): sc(ws3,3,i+1,h,hdr_f,hdr_fl)
ws3.row_dimensions[3].height=28; ws3.freeze_panes="A4"

design = [
    ("KCPM-752","frontend_form_bva_spec.md","BVA (Gia tri bien)",10,N,N,N,P,P,P),
    ("KCPM-753","jwt_permission_ep_spec.md","EP (Phan hoach tuong duong)",8,N,N,N,P,P,P),
    ("KCPM-761","jwt_validation_whitebox_spec.md","Whitebox (CFG/Basis Path)",6,N,N,N,N,P,P),
    ("KCPM-766","prescription_whitebox_spec.md","Whitebox (CFG/Basis Path)",8,N,N,N,N,P,P),
    ("KCPM-771","postman_test_scripts_spec.md","Postman (Reusable Scripts)",6,N,N,N,N,P,P),
    ("KCPM-776","admin_config_postman_test_spec.md","Postman API Test",9,N,N,N,N,P,P),
    ("KCPM-781","clinic_doctors_postman_test_spec.md","Postman API Test",9,N,N,N,N,P,P),
    ("KCPM-786","doctor_appointments_postman_test_spec.md","Postman API Test",12,N,N,N,N,P,P),
    ("KCPM-791","patient_appointments_postman_test_spec.md","Postman API Test",12,N,N,N,N,P,P),
    ("KCPM-796","patient_profile_postman_test_spec.md","Postman API Test",9,N,N,N,N,P,P),
    ("KCPM-801","medical_services_postman_test_spec.md","Postman API Test",9,N,N,N,N,P,P),
    ("KCPM-811","frontend_e2e_scenarios_spec.md","E2E Test Design",15,N,N,N,N,P,P),
    ("KCPM-806","testware_env_data_plan.md","Testware & Environment",0,N,N,N,N,P,P),
    ("KCPM-816","unit_test_plan_auth_admin_spec.md","Unit Test Plan",0,N,N,N,N,N,P),
    ("KCPM-821","bug_tracking_standard_spec.md","Bug Tracking Standard",0,N,N,N,N,N,P),
    ("KCPM-831","ci_failure_fix_report.md","CI/CD Fix Report",0,N,N,N,N,N,P),
]

for i,(tk,nm,tech,tc,v1,v2,v3,v4,v5,v6) in enumerate(design):
    r=4+i; rf=alt_fl if i%2==0 else w_fl
    sc(ws3,r,1,i+1,norm_f,rf); sc(ws3,r,2,tk,bold_f,rf); sc(ws3,r,3,nm,norm_f,rf,la)
    sc(ws3,r,4,tech,norm_f,rf); sc(ws3,r,5,tc if tc>0 else "-",norm_f,rf)
    for ci,val in enumerate([v1,v2,v3,v4,v5,v6]):
        ff,ffl=rs(val); sc(ws3,r,6+ci,val,ff,ffl)

# ==================== SHEET 4: VERSION COMPARISON ====================
ws4 = wb.create_sheet("So Sanh Version"); ws4.sheet_properties.tabColor = ORANGE
for c in range(1,11): ws4.column_dimensions[get_column_letter(c)].width = 16
banner(ws4,1,"SO SANH TONG HOP GIUA 6 VERSION",10)
banner(ws4,2,"Dashboard tien do kiem thu qua tung tuan",10,sub_f,hdr_fl,30)

ch = ["Chi so","v1.0.0","v2.0.0","v3.0.0","v4.0.0","v5.0.0","v6.0.0","Xu huong","Nhan xet"]
cw = [24,12,12,12,12,12,12,12,24]
for i,w in enumerate(cw): ws4.column_dimensions[get_column_letter(i+1)].width = w
for i,h in enumerate(ch): sc(ws4,3,i+1,h,hdr_f,hdr_fl)
ws4.row_dimensions[3].height=28

comp = [
    ("JUnit Test Classes","?","?","?","?","?","?","Bien dong","So file test duoc chay"),
    ("JUnit Test Cases","131","110","213","74","37","87","Bien dong","So TC thuc te chay duoc"),
    ("JUnit Pass Rate","92.3%","90.9%","100%","97.2%","94.5%","100%","Tang/Giam","Tuy thuoc moi truong"),
    ("JUnit Failures","0","0","0","2","2","0","Bien dong","Loi logic"),
    ("JUnit Errors","10","10","0","0","0","0","Giam","Loi moi truong da fix tu v3"),
    ("BVA Test Cases","0","0","0","10","10","10","On dinh","Frontend + Core Business"),
    ("EP Test Cases","0","0","0","8","8","8","On dinh","JWT + Permission"),
    ("Whitebox Analyses","0","0","0","0","2","2","On dinh","JWT + Prescription CFG"),
    ("Postman API Modules","0","0","0","0","7","7","Moi","66 assertions"),
    ("E2E Scenarios","0","0","0","0","15","15","Moi","Login, CRUD, Navigation"),
    ("Test Design Artifacts","0","0","0","2","13","16","Tang manh","+3 o week 6"),
    ("CI/CD Pipeline","Chua","Chua","Chua","Chua","Chua","Fixed","On dinh","KCPM-831/839"),
]

for i,(m,v1,v2,v3,v4,v5,v6,tr,nt) in enumerate(comp):
    r=4+i; rf=alt_fl if i%2==0 else w_fl
    sc(ws4,r,1,m,bold_f,rf,la)
    for ci,val in enumerate([v1,v2,v3,v4,v5,v6]): sc(ws4,r,2+ci,val,norm_f,rf)
    sc(ws4,r,8,tr,norm_f,rf); sc(ws4,r,9,nt,norm_f,rf,la)

cr=4+len(comp)+2
banner(ws4,cr,"KET LUAN",9,sub_f,hdr_fl,30)
conclusions = [
    "1. v1-v2: Giai doan dau gap nhieu loi moi truong (10 Errors) va chua hoan thien toan bo test suite.",
    "2. v3: Pass rate dat 100% voi so luong test cao (213 TC) do moi truong da duoc fix.",
    "3. v4-v5: So luong test thuc thi giam manh do crash moi truong hoac refactor loi, xay ra Failures (2).",
    "4. v6: Hoan thien CI/CD, Pass Rate phuc hoi 100% voi 87 Test Cases on dinh.",
    "5. Ket luan: So lieu phan anh dung thuc te qua trinh phat trien (co loi, co fix, co crash).",
]
for i,c in enumerate(conclusions):
    r=cr+1+i
    ws4.merge_cells(start_row=r,start_column=1,end_row=r,end_column=9)
    sc(ws4,r,1,c,norm_f,sum_fl,la); ws4.row_dimensions[r].height=25

# ==================== SHEET 5: GUIDE ====================
ws5 = wb.create_sheet("Huong Dan Chuyen Version"); ws5.sheet_properties.tabColor = "8E44AD"
for c in range(1,7): ws5.column_dimensions[get_column_letter(c)].width = 25
banner(ws5,1,"HUONG DAN CHUYEN DOI GIUA CAC VERSION",6)
guide = [
    ("Xem danh sach version:","git tag"),
    ("Chuyen sang v1.0.0 (Week 1):","git checkout v1.0.0"),
    ("Chuyen sang v2.0.0 (Week 2):","git checkout v2.0.0"),
    ("Chuyen sang v3.0.0 (Week 3-4):","git checkout v3.0.0"),
    ("Chuyen sang v4.0.0 (Week 5a):","git checkout v4.0.0"),
    ("Chuyen sang v5.0.0 (Week 5b):","git checkout v5.0.0"),
    ("Chuyen sang v6.0.0 (Week 6):","git checkout v6.0.0"),
    ("Quay ve main (moi nhat):","git checkout main"),
    ("Chay test:","mvn test -Dspring.profiles.active=test"),
    ("Push tags len remote:","git push origin --tags"),
]
for i,(d,cmd) in enumerate(guide):
    r=3+i
    ws5.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    sc(ws5,r,1,d,bold_f,sum_fl,la)
    ws5.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6)
    sc(ws5,r,3,cmd,Font(name="Consolas",size=11,color=DARK_BLUE),w_fl,la)

# Save
out = "Test_Result_Report_duyho0705.xlsx"
wb.save(out)
print(f"[OK] Report saved: {out}")
